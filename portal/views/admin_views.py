import csv
import openpyxl
from django.shortcuts import render, redirect
from django.http import HttpResponse, JsonResponse
from portal.models import NGUOIDUNG
from django.views.decorators.csrf import csrf_protect, csrf_exempt
from django.core import serializers
from portal.models import *
import hashlib
import requests
import math
from django.core.serializers.json import DjangoJSONEncoder
import json
from . import chucnang as ChucNang
MoiTrang = 10
def kiemTraCookie(request):
    print(request)
def index(request):
    return redirect('sinhvien/')
@csrf_exempt
def xoa_nguoidung(request, nguoidungID):
      if (request.method== "POST"):
            try:
                NGUOIDUNG.objects.filter(pk=nguoidungID).update(HoatDong=False)  
                return HttpResponse(json.dumps({'code': 200, 'msg': 'success'}))
            except:
                return HttpResponse(json.dumps({'code': 403, 'msg': 'Không thể xóa người dùng'}))
      return HttpResponse(json.dumps({'code': 403, 'msg': 'Not allow method'}))
@csrf_exempt
def them_nguoidung(request):
    # Giống hệt đăng ký
    if request.method == "GET":
        return render(request, 'portal/admin/them_nguoidung.html')
    if request.method == "POST":
        # Lấy dữ liệu đăng ký dưới dạng Objects
        thongTinTao = json.loads(request.body)
        # 'SDT': '123', 'Email': '1234@gmail.com', 'GioiTinh': '0', 'NgaySinh': '2020-12-10', 'MatKhau': 'password', 'Quyen': 3}
        maHoaMatKhau = hashlib.md5(thongTinTao.get(
            'MatKhau').encode()).hexdigest()  # Mã hóa mật khẩu md5
        # Các thao tác validate ở đây
        try:
            nguoiDungDk = NGUOIDUNG(TenNguoiDung=thongTinTao.get('TenNguoiDung'), HoTen=thongTinTao.get('HoTen'), SDT=thongTinTao.get(
                'SDT'), MatKhau=maHoaMatKhau, Email=thongTinTao.get('Email'), NgaySinh=thongTinTao.get('NgaySinh'), GioiTinh=thongTinTao.get('GioiTinh'), Quyen=thongTinTao.get('Quyen'))
            nguoiDungDk.save()
        except Exception as insertErr:
            resp = {"code": 404}
            resp['msg'] = str(insertErr)
            return HttpResponse(json.dumps(resp))
        resp = {"code": 200}
        resp['msg'] = "success"
        return HttpResponse(json.dumps(resp))
    return HttpResponse(json.dumps({'code': 403, 'msg': 'Not allow method'}))
@csrf_exempt
def sua_nguoidung(request, nguoidungID):
    # Giống hệt đăng ký
    if request.method == "GET":
        thongTinNguoiDungList = NGUOIDUNG.objects.filter(pk=nguoidungID).values()
        if (len(thongTinNguoiDungList)< 1):
            return HttpResponse(json.dumps({'code': 404, 'msg': 'User not found'}))
        thongTinNguoiDung = thongTinNguoiDungList[0]
        print(thongTinNguoiDung)
        return render(request, 'portal/admin/sua_nguoidung.html', thongTinNguoiDung)
    if request.method == "POST":
        thongTinUpdate = json.loads(request.body)
        try:
            NGUOIDUNG.objects.filter(pk=nguoidungID).update(Email =thongTinUpdate.get('Email') ,GioiTinh= thongTinUpdate.get('GioiTinh') ,HoTen= thongTinUpdate.get('HoTen') ,NgaySinh= thongTinUpdate.get('NgaySinh') ,Quyen= thongTinUpdate.get('Quyen') ,SDT= thongTinUpdate.get('SDT'))  
            if (thongTinUpdate.get('MatKhau')!=""):
                NGUOIDUNG.objects.filter(pk=nguoidungID).update(MatKhau = hashlib.md5(thongTinUpdate.get('MatKhau').encode()).hexdigest()) # Mã hóa mật khẩu md5
        except Exception as insertErr:
            resp = {"code": 404}
            resp['msg'] = str(insertErr)
            return HttpResponse(json.dumps(resp))
        resp = {"code": 200}
        resp['msg'] = "success"
        return HttpResponse(json.dumps(resp))
    return HttpResponse(json.dumps({'code': 403, 'msg': 'Not allow method'}))
def quanly_sinhvien(request, trang=1):  # Mặc định trang = 1
    trangHienTai = int(trang)
    # Lấy danh sách sinh viên theo trang / mỗi trang
    dsSinhVien_phanTrang = danhSachNguoiDung(trangHienTai, MoiTrang, 3)
    # Tạo JSON để render --> chứa phân trang, ds người dùng
    content = taoJsonQLNguoiDung(dsSinhVien_phanTrang, trangHienTai, 'sinhvien')
    return render(request, 'portal/admin/ql_nguoidung.html', content)
def quanly_giangvien(request, trang=1):  # Mặc định trang = 1
    trangHienTai = int(trang)
    # Lấy danh sách sinh viên theo trang / mỗi trang
    dsGiangVien_phanTrang = danhSachNguoiDung(trangHienTai, MoiTrang, 2)
    # Tạo JSON để render --> chứa phân trang, ds người dùng
    content = taoJsonQLNguoiDung(dsGiangVien_phanTrang, trangHienTai, 'giangvien')
    return render(request, 'portal/admin/ql_nguoidung.html', content)
def quanly_thongbao(request, trang=1):  # Mặc định trang = 1
    trangHienTai = int(trang)
    # Lấy danh sách sinh viên theo trang / mỗi trang
    dsThongBao_phanTrang = danhSachThongBao(trangHienTai, MoiTrang)
    # Tạo JSON để render --> chứa phân trang, ds người dùng
    content = taoJsonQLNguoiDung(dsThongBao_phanTrang, trangHienTai, 'thongbao')
    return render(request, 'portal/admin/ql_thongbao.html', content)
def quanly_hoatdong(request):  # Mặc định trang = 1
    sql = """SELECT
        portal_hoatdong.IdHoatDong,
        portal_hoatdong.IdUser,
        portal_hoatdong.ChiTiet,
        portal_hoatdong.NgayBD,
        portal_hoatdong.NgayKT,
        portal_hoatdong.SoLuong,
        portal_hoatdong.DiemRL,
        portal_hoatdong.HoatDong,
        portal_hoatdong.DaDangKi,
        portal_hoatdong.DangThucHien,
        portal_hoatdong.TenHoatDong,
        portal_hoatdong.Ki,
        portal_hoatdong.Loai,
        portal_nguoidung.HoTen,
        portal_loaihd.TenLoaiHD
        FROM
        portal_hoatdong
        JOIN portal_nguoidung ON portal_hoatdong.IdUser = portal_nguoidung.IdUser
        JOIN portal_loaihd ON portal_loaihd.idLoaiHD = portal_hoatdong.Loai
        WHERE portal_hoatdong.HoatDong = 1
        """
    dsHoatDong = ChucNang.TruyVanDuLieu(sql)
    content = {
        "DS_NguoiDung" : dsHoatDong['data']
    }
    return render(request, 'portal/admin/ql_hoatdong.html', content)
def danhSachThongBao(trang, moitrang,search=""):
    dsDetai = tatCaThongBao(search)
    tongDetai = len(dsDetai['data'])
    # Tổng số trang = tổng sinh viên / số sV mỗi trang , làm tròn lên
    tongTrang = math.ceil(tongDetai / moitrang)
    if (trang > tongTrang):
        trang = tongTrang  # VD : tổng 4 trang, yêu cầu trang 4 --> chỉ load tới trang 3
    if (trang < 1):
        trang = 1  # VD : tổng 4 trang, yêu cầu trang 4 --> chỉ load tới trang 3
    sql = """SELECT * from portal_thongbao where portal_thongbao.ChiTiet LIKE '%{0}%' OR portal_thongbao.TieuDe LIKE '%{1}%'
            LIMIT {2} OFFSET {3}
            """.format(search,search,moitrang, (trang-1)*moitrang)  # Offset bắt đầu từ 0 --> trang - 1, công thức phân trang sql
    dsDeTai_PhanTrang = ChucNang.TruyVanDuLieu(sql)
    dsDeTai_PhanTrang['SoDeTai'] = len(dsDetai['data'])
    dsDeTai_PhanTrang['SoTrang'] = tongTrang
    return dsDeTai_PhanTrang
def tatCaThongBao(search=""):
    danhSachDeTai = ChucNang.TruyVanDuLieu("""SELECT * from portal_thongbao where portal_thongbao.ChiTiet LIKE '%{0}%' OR portal_thongbao.TieuDe LIKE '%{1}%'""".format(search,search))
    return danhSachDeTai
# Hàm logic
def danhSachNguoiDung(trang, moitrang, quyen):
    dsNguoiDung = tatCaNguoiDung(quyen)
    tongNguoiDung = len(dsNguoiDung['data'])
    # Tổng số trang = tổng sinh viên / số sV mỗi trang , làm tròn lên
    tongTrang = math.ceil(tongNguoiDung / moitrang)
    if (trang > tongTrang):
        trang = tongTrang  # VD : tổng 4 trang, yêu cầu trang 4 --> chỉ load tới trang 3
    if (trang < 1):
        trang = 1  # VD : tổng 4 trang, yêu cầu trang 4 --> chỉ load tới trang 3
    sql = "SELECT * FROM `portal_nguoidung` WHERE Quyen = {0} AND HoatDong = 1 LIMIT {1} OFFSET {2}".format(
        quyen, moitrang, (trang-1)*moitrang)  # Offset bắt đầu từ 0 --> trang - 1, công thức phân trang sql
    dsNguoiDung_phanTrang = querySetToJson(NGUOIDUNG.objects.raw(sql))
    dsNguoiDung_phanTrang['SoNguoiDung'] = len(dsNguoiDung['data'])
    dsNguoiDung_phanTrang['SoTrang'] = tongTrang
    return dsNguoiDung_phanTrang
def querySetToJson(rawquerySet):
    # Chuyển rawQuerySet thành dạng Json với các phần tử là các fields
    data = serializers.serialize('json', rawquerySet)
    listData = json.loads(data)  # Chuyển sang JSON
    jsonData = {}
    mangPhanTu = []
    for phanTu in listData:
        # Với mỗi field --> lấy đoạn json chứa thông tin (rows) của tables, lưu vào mảng
        try:
            phanTu['fields']['id'] = phanTu['pk']
            # thêm id vô fields
        except:
            pass
        mangPhanTu.append(phanTu['fields'])
    jsonData['data'] = mangPhanTu  # Lưu vào JSON để trả về
    return jsonData  # Trả về  JSON
    # return json.dumps(jsonData)   #Trả về string JSON
def tatCaNguoiDung(quyen):
    danhSachNguoiDung = querySetToJson(NGUOIDUNG.objects.raw(
        'SELECT * FROM `portal_nguoidung` WHERE Quyen = {0}  AND HoatDong = 1'.format(quyen)))
    return danhSachNguoiDung
# Hàm tạo dữ liệu json để render vào quanli_nguoidung.html (dùng chung cho cả 3 loại người dùng)
def taoJsonQLNguoiDung(dsNguoiDung, trangHienTai, loaiNguoiDung):
    tongSoTrang = int(dsNguoiDung['SoTrang'])  # Lấy tổng số trang
    if (tongSoTrang ==0):
              tongSoTrang = 1
    if trangHienTai > tongSoTrang:  # Nếu trang yêu cầu > tổng số --> Quay về trang 1
        return redirect('/')
    # Chỗ này hiển thị mấy nút đến trang  1 2 3 4 5 Cuối
    dsTrang = range(trangHienTai - 5, trangHienTai + 5)
    if (trangHienTai-5 < 1):
        dsTrang = range(1, 11)
    if (trangHienTai + 5 > tongSoTrang):
        dsTrang = range(tongSoTrang-10, tongSoTrang+1)
    if (tongSoTrang < 10):
        dsTrang = range(1, tongSoTrang+1)
    trangSau = trangHienTai + 1
    if trangHienTai > tongSoTrang:
        trangHienTai = tongSoTrang
    trangTruoc = trangHienTai - 1
    if trangTruoc < 1:
        trangTruoc = 1
    # Tạo Json để render ra HTML
    tenLoai = ""
    if loaiNguoiDung == 'admin':
        tenLoai = "admin"
    if loaiNguoiDung == 'giangvien':
        tenLoai = "giảng viên"
    if loaiNguoiDung == 'sinhvien':
              tenLoai = "sinh viên"
    if loaiNguoiDung == 'thongbao':
              tenLoai = "thông báo"
    tongso = 0
    try:
        tongso = dsNguoiDung['SoNguoiDung']
    except:
        tongso = dsNguoiDung['SoDeTai']
    content = {
        'NguoiDung': loaiNguoiDung,
        'LoaiNguoiDung': tenLoai,
        'DS_NguoiDung': dsNguoiDung['data'],
        'TongNguoiDung': tongso,
        'SoTrang': dsTrang,
        'TrangHienTai': trangHienTai,
        'TongTrang': tongSoTrang,
        'SoNguoiDung': len(dsNguoiDung['data']),
        'TrangSau':  trangSau,
        'TrangTruoc': trangTruoc
    }
    return content
def dieukienexportDetai(row):
    sql = "SELECT * from portal_nguoidung WHERE IdUser = {0}".format(row[1].value)
    if (len(ChucNang.TruyVanDuLieu(sql)['data'])==0):
        return False
    sql = "SELECT * from portal_loaidetai WHERE IdLoai = {0}".format(row[6].value)
    if (len(ChucNang.TruyVanDuLieu(sql)['data'])==0):
        return False
    sql = "SELECT * from portal_khoa WHERE IdKhoa = {0}".format(row[7].value)
    if (len(ChucNang.TruyVanDuLieu(sql)['data'])==0):
        return False 
    return True
@csrf_exempt
def importdetai(request):
    if request.method == "POST":
        if (len(request.FILES) < 1):
            return HttpResponse(json.dumps({'code': 403, 'msg': 'No file'}))
        sql = ""
        try:
            excel_file = request.FILES["file"]
            wb = openpyxl.load_workbook(excel_file)
            worksheet = wb.worksheets[0]
            excel_data = list()
            count = 0
            for row in worksheet.iter_rows():
                if (count == 0):
                    count = 1
                    continue
                if (dieukienexportDetai(row)):
                    sql = ""
                    checksql = "SELECT IdDetai FROM portal_detai WHERE IdDetai={0} LIMIT 1 ".format(row[0].value) #Kiemtra ID De TaI
                    # Kiemtra ton tai 
                    temp = ChucNang.TruyVanDuLieu(checksql)
                    if (len(temp['data']) == 1):
                        sql = """
                        UPDATE `portal`.`portal_detai` SET `IdUser` = {1}, `ChiTiet` = '{2}', `NgayBD` = '{3}', `NgayKT` = '{4}', `SoLuong` = {5}, `IdLoai` = {6},`TenDeTai` = '{7}', `IdKhoa` = {8} WHERE `IdDeTai` = {0}
                        """.format(row[0].value,row[1].value,row[2].value,row[3].value,row[4].value,row[5].value,row[6].value,row[8].value,row[7].value)
                    else:
                        sql = """ 
                        INSERT INTO `portal`.`portal_detai`(`IdUser`,`ChiTiet`, `NgayBD`, `NgayKT`, `SoLuong`, `IdLoai`, `TenDeTai`, `IdKhoa`) VALUES ({0},'{1}', '{2}', '{3}', {4}, {5}, '{6}', {7})
                        """.format(row[1].value,row[2].value,row[3].value,row[4].value,row[5].value,row[6].value,row[8].value,row[7].value)
                    ChucNang.UpdateDuLieu(sql)
        except Exception as insertErr:
            resp = {"code": 404}
            resp['msg'] = str(insertErr)
            return HttpResponse(json.dumps(resp))
        resp = {"code": 200}
        resp['msg'] = "success"
        return HttpResponse(json.dumps(resp))
    return HttpResponse(json.dumps({'code': 403, 'msg': 'Not allow method'}))
@csrf_exempt
def importExcel(request, loai):
    if request.method == "POST":
        if (len(request.FILES) < 1):
            return HttpResponse(json.dumps({'code': 403, 'msg': 'No file'}))
        
        try:
            excel_file = request.FILES["file"]
            wb = openpyxl.load_workbook(excel_file)
            worksheet = wb.worksheets[0]
            excel_data = list()
            count = 0
            for row in worksheet.iter_rows():
                if (count == 0):
                    count = 1
                    continue
                sql = ""
                checksql = "SELECT IdUser FROM portal_nguoidung WHERE TenNguoiDung='{0}' LIMIT 1 ".format(row[1].value)
                temp = ChucNang.TruyVanDuLieu(checksql)
                if (len(temp['data']) == 1):
                    sql = """
                    UPDATE portal_nguoidung SET HoTen = '{0}', Email= '{1}' ,SDT= '{2}',    GioiTinh= {3},  Khoa= '{4}',     NgaySinh = '{5}', HoatDong = {6} WHERE TenNguoiDung='{7}'
                    """.format(                 row[2].value,   row[3].value,row[4].value,  row[7].value,   row[5].value,       row[6].value,  row[8].value,        row[1].value)
                else:
                    sql = """ 
                    INSERT INTO portal_nguoidung(HoTen,MatKhau,Email,Quyen,SDT,NgaySinh,GioiTinh,TenNguoiDung,HoatDong,Khoa) VALUES ('{0}', '{1}', '{2}', {3}, '{4}', '{5}', {6}, '{7}', 1, {8})
                    """.format(row[2].value, row[1].value, row[3].value, 3, row[4].value,  row[6].value, row[7].value, row[1].value,1,  row[8].value)
                ChucNang.UpdateDuLieu(sql)
                checksql = "SELECT * FROM portal_diemtrungbinh WHERE userID='{0}' LIMIT 1 ".format(row[1].value)
                temp = ChucNang.TruyVanDuLieu(checksql)
                if (len(temp['data']) == 1):
                          sql = """
                    UPDATE portal_diemtrungbinh SET Diem = '{0}' WHERE userID='{1}'
                    """.format(row[9].value, row[1].value)
                else:
                    sql = """ 
                    INSERT INTO portal_diemtrungbinh(userID,Diem) VALUES ('{0}', {1})
                    """.format(row[0].value,row[9].value)
                ChucNang.UpdateDuLieu(sql)
                # Id,  1                  2             3         4             5         6                  7            8         9
                # Id,  TenNguoiDung     HoTen       Email       SDT         Khoa         NgaySinh           GioiTinh  HoatDong   Diem
            #    ['2', '2033172027', 'Võ Phú Hải', 'phuhai113@gmail.com', 'None', '08', '1998-10-10 00:00:00', 'None', '1', '9.33']
        except Exception as insertErr:
            resp = {"code": 404}
            resp['msg'] = str(insertErr)
            return HttpResponse(json.dumps(resp))
        resp = {"code": 200}
        resp['msg'] = "success"
        return HttpResponse(json.dumps(resp))
    return HttpResponse(json.dumps({'code': 403, 'msg': 'Not allow method'}))
# Chi tiet thong bao 
def chitiet_thongbao(request, idTB):
    sql = "SELECT * FROM portal_thongbao WHERE IdThongBao={0}".format(idTB)
    data = ChucNang.TruyVanDuLieu(sql)
    chitiet = data['data'][0]
    return render(request, 'portal/admin/chitiet_thongbao.html', {'ThongBao':chitiet})
# Chi tiet thong bao 
@csrf_exempt
def sua_thongbao(request, idTB):
    if request.method == "GET":
        sql = "SELECT * FROM portal_thongbao WHERE IdThongBao={0}".format(idTB)
        data = ChucNang.TruyVanDuLieu(sql)
        chitiet = data['data'][0]
        return render(request, 'portal/admin/sua_thongbao.html', {'ThongBao':chitiet})
    else:
        tenHoatDong = request.POST['tenThongBao']
        Ngay = request.POST['ngayKTHD']
        ChiTiet = request.POST['chiTietHD']
        jsonRender = {'tieude' : 'Thành công', 'ThongBao' : 'Thành Công','ChiTiet' : 'Sửa thông báo thành công', 'backlink': '/admin/sua_thongbao/{0}'.format(idTB)}
        updatesql = "UPDATE portal_thongbao SET ChiTiet  = '{0}',  TieuDe = '{1}',  NgayThongBao = '{2}' WHERE  IdThongBao  = {3}".format(ChiTiet,tenHoatDong, Ngay,idTB)
        try:
            ChucNang.UpdateDuLieu(updatesql)
        except Exception as exc:
            jsonRender['ChiTiet'] = str(exc)
            jsonRender['ThongBao'] = str("Không thành công")
        return render(request, 'portal/giangvien/thongbao.html', jsonRender)
@csrf_exempt
def chitiethoatdong(request, id):
    if request.method == "GET":
        sqlChiTietHD = """SELECT
        portal_hoatdong.IdHoatDong,
        portal_hoatdong.IdUser,
        portal_hoatdong.ChiTiet,
        portal_hoatdong.NgayBD,
        portal_hoatdong.NgayKT,
        portal_hoatdong.SoLuong,
        portal_hoatdong.DiemRL,
        portal_hoatdong.HoatDong,
        portal_hoatdong.DaDangKi,
        portal_hoatdong.DangThucHien,
        portal_hoatdong.TenHoatDong,
        portal_hoatdong.Ki,
        portal_hoatdong.Loai,
        portal_nguoidung.HoTen,
        portal_loaihd.TenLoaiHD
        FROM
        portal_hoatdong
        JOIN portal_nguoidung ON portal_hoatdong.IdUser = portal_nguoidung.IdUser
        JOIN portal_loaihd ON portal_loaihd.idLoaiHD = portal_hoatdong.Loai
        WHERE portal_hoatdong.HoatDong = 1 AND portal_nguoidung.HoatDong = 1 AND portal_hoatdong.IdHoatDong = {0}
        """.format(id)
        chiTietHoatDong_qr = ChucNang.TruyVanDuLieu(sqlChiTietHD)
        if (len(chiTietHoatDong_qr['data'])==0):
            jsonRender = {'tieude' : 'Không thành công', 'ThongBao' : 'Không thành công','ChiTiet' : 'Không tìm thấy hoạt động', 'backlink': '/admin/dshoatdong'}
            return render(request, 'portal/giangvien/thongbao.html', jsonRender)
        sql = """ SELECT
                            portal_hoatdongdadangky.IdHDDDK,
                            portal_hoatdongdadangky.IdHoatDong,
                            portal_hoatdongdadangky.DaChamDiem,
                            portal_nguoidung.IdUser,
                            portal_nguoidung.TenNguoiDung,
                            portal_nguoidung.HoTen,
                            portal_hoatdong.DiemRL,
                            portal_hoatdong.Ki

                            FROM
                                portal_hoatdongdadangky
                                JOIN portal_nguoidung ON portal_nguoidung.IdUser = portal_hoatdongdadangky.IdUser
                                JOIN portal_hoatdong ON portal_hoatdong.IdHoatDong = portal_hoatdongdadangky.IdHoatDong
                            WHERE
                            portal_hoatdongdadangky.IdHoatDong = {0} AND
                            portal_hoatdong.HoatDong = 1 AND
                            portal_nguoidung.HoatDong = 1
        """.format(id)
        dsSVThamGia = ChucNang.TruyVanDuLieu(sql)
        chitiet = chiTietHoatDong_qr['data'][0]
        return render(request, 'portal/admin/chitiethoatdong.html', {'HoatDong':chitiet, 'DSThamGia' : dsSVThamGia['data']})
@csrf_exempt
def sua_thongbao(request, idTB):
    if request.method == "GET":
        sql = "SELECT * FROM portal_thongbao WHERE IdThongBao={0}".format(idTB)
        data = ChucNang.TruyVanDuLieu(sql)
        chitiet = data['data'][0]
        return render(request, 'portal/admin/sua_thongbao.html', {'ThongBao':chitiet})
    else:
        tenHoatDong = request.POST['tenThongBao']
        Ngay = request.POST['ngayKTHD']
        ChiTiet = request.POST['chiTietHD']
        jsonRender = {'tieude' : 'Thành công', 'ThongBao' : 'Thành Công','ChiTiet' : 'Sửa thông báo thành công', 'backlink': '/admin/sua_thongbao/{0}'.format(idTB)}
        updatesql = "UPDATE portal_thongbao SET ChiTiet  = '{0}',  TieuDe = '{1}',  NgayThongBao = '{2}' WHERE  IdThongBao  = {3}".format(ChiTiet,tenHoatDong, Ngay,idTB)
        try:
            ChucNang.UpdateDuLieu(updatesql)
        except Exception as exc:
            jsonRender['ChiTiet'] = str(exc)
            jsonRender['ThongBao'] = str("Không thành công")
        return render(request, 'portal/giangvien/thongbao.html', jsonRender)
@csrf_exempt
def diemdanh(request):
    if request.method == "GET":
        jsonRender = {'tieude' : 'Thành công', 'ThongBao' : 'KHONG CO PHEP','ChiTiet' : 'POST REQUIRED', 'backlink': '/admin/dshoatdong'}
        return render(request, 'portal/giangvien/thongbao.html', jsonRender)
    else:
        print(request)
        dsSV = json.loads(request.body)
        resp = {"code": 200}
        resp['msg'] = "success"
        for svdk in dsSV:
            try:
                sql = "UPDATE `portal`.`portal_hoatdongdadangky` SET `DaChamDiem` = {0} WHERE `IdHoatDong` = {1} AND IdUser={2}".format(svdk['dadiemdanh'], svdk['hoatdongid'], svdk['id'])
                ChucNang.UpdateDuLieu(sql)
                if (svdk['dadiemdanh']):
                    checkTonTaiSql = "SELECT * from portal_diemrenluyen where hoatdongid = {0} and userID = {1} and Ki={2}".format(svdk['hoatdongid'], svdk['id'], svdk['ki'])
                    tontai_res = ChucNang.TruyVanDuLieu(checkTonTaiSql)
                    if (len(tontai_res['data'])==0): #SV chưa có điểm
                        themdiemSql = "INSERT INTO portal_diemrenluyen(userID,Diem,Ki,hoatDongId) VALUES ({0},{1},{2},{3})".format(svdk['id'],svdk['diem'],svdk['ki'], svdk['hoatdongid'])
                        ChucNang.UpdateDuLieu(themdiemSql)
                else:
                    checkTonTaiSql = "SELECT * from portal_diemrenluyen where hoatdongid = {0} and userID = {1} and Ki={2}".format(svdk['hoatdongid'], svdk['id'], svdk['ki'])
                    tontai_res = ChucNang.TruyVanDuLieu(checkTonTaiSql)
                    if (len(tontai_res['data'])!=0): #SV chưa có điểm
                        deleteSql = "DELETE FROM portal_diemrenluyen WHERE hoatdongid = {0} AND userID={1} and Ki={2}".format(svdk['hoatdongid'], svdk['id'], svdk['ki'])
                        ChucNang.UpdateDuLieu(deleteSql)
            except Exception as errr:
                resp['msg'] = str(errr)
                return HttpResponse(json.dumps(resp))
        return HttpResponse(json.dumps(resp))
def exportHoatDong(request, id):
    if request.method == "GET":
        sql ="""
                     SELECT
                portal_hoatdongdadangky.IdHoatDong,
                portal_hoatdongdadangky.IdUser,
                portal_nguoidung.TenNguoiDung,
                portal_nguoidung.HoTen,
                portal_nguoidung.Khoa,
                portal_nguoidung.SDT,
                portal_nguoidung.Email,
                portal_hoatdongdadangky.NgayDKHD,
                hoatdong_nguoidung.HoTen as NguoiDH,
                hoatdong_nguoidung.TenHoatDong,
                hoatdong_nguoidung.Ki,
                portal_hoatdongdadangky.DaChamDiem
                FROM
                portal_hoatdongdadangky
                JOIN portal_nguoidung ON portal_nguoidung.IdUser = portal_hoatdongdadangky.IdUser
                JOIN (SELECT HoTen,DiemRL, Ki,IdHoatDong, TenHoatDong FROM portal_hoatdong JOIN portal_nguoidung on portal_nguoidung.IdUser = portal_hoatdong.IdUser) AS hoatdong_nguoidung ON hoatdong_nguoidung.IdHoatDong= portal_hoatdongdadangky.IdHoatDong
            WHERE portal_hoatdongdadangky.IdHoatDong = {0}
        """.format(id)
        dataexport_res = ChucNang.TruyVanDuLieu(sql)
        header = 0
        try:
            with open('csvfile.csv', mode='w',newline='', encoding='UTF-8') as csvfile:
                for item in dataexport_res['data']:
                    csvfile_writer_writer = csv.writer(csvfile, delimiter=',')
                    if (header ==0):
                        csvfile_writer_writer.writerow(['IDHoatDong', 'IdUser', 'MSSV', 'HoTen', 'Khoa', 'SDT', 'Email', 'NgayDKHD', 'NguoiDH', 'TenHoatDong', 'Ki', 'DaDiemDanh'])
                        header = 1
                    csvfile_writer_writer.writerow([str(item['IdHoatDong']), str(item['IdUser']), str(item['TenNguoiDung']), str(item['HoTen']), str(item['Khoa']), str(item['SDT']),str(item['Email']), item['NgayDKHD'].strftime("%m/%d/%Y %H:%M:%S"), str(item['NguoiDH']), str(item['TenHoatDong']), str(item['Ki']), str(item['DaChamDiem'])])
            with open('csvfile.csv') as myfile:
                response = HttpResponse(myfile, content_type='text/csv')
                response['Content-Disposition'] = 'attachment; filename=csvfile.csv'
                return response
        except  Exception as err:
            print(err)
def exportdkdt(request):
    if request.method == "GET":
        sql ="""
            SELECT
                portal_detaidadangky.IdDTDDK,
                portal_detaidadangky.IdUser,
                portal_detaidadangky.NgayDKDT,
                portal_nguoidung.HoTen,
                portal_nguoidung.TenNguoiDung as MSSV,
                detai_giangvien.TenGiangVien,
                detai_giangvien.IDGV,
                portal_detaidadangky.IdDeTai,
                detai_giangvien.TenDeTai
            FROM
                portal_detaidadangky
                JOIN portal_nguoidung ON portal_nguoidung.IdUser = portal_detaidadangky.IdUser
                JOIN (SELECT TenDeTai, HoTen as TenGiangVien, portal_nguoidung.IdUser as IDGV,portal_detai.IdDeTai FROM portal_detai join portal_nguoidung  on portal_nguoidung.IdUser = portal_detai.IdUser) as detai_giangvien on detai_giangvien.IdDeTai =  portal_detaidadangky.IdDeTai
        """
        dataexport_res = ChucNang.TruyVanDuLieu(sql)
        header = 0
        try:
            with open('DanhSachDKDT.csv', mode='w',newline='', encoding='UTF-8') as csvfile:
                for item in dataexport_res['data']:
                    csvfile_writer_writer = csv.writer(csvfile, delimiter=',')
                    if (header ==0):
                            #                           IdDTDDK	    IdUser	  NgayDKDT	  HoTen	  MSSV	TenGiangVien	IDGV	IdDeTai	TenDeTai
                        csvfile_writer_writer.writerow(['IdDTDDK', 'IdUser', 'NgayDKDT', 'HoTen', 'MSSV', 'TenGiangVien', 'IDGV', 'IdDeTai', 'TenDeTai'])
                        header = 1
                    csvfile_writer_writer.writerow([str(item['IdDTDDK']), str(item['IdUser']), str(item['NgayDKDT'].strftime("%m/%d/%Y %H:%M:%S")), str(item['HoTen']), str(item['MSSV']), str(item['TenGiangVien']),str(item['IDGV']), item['IdDeTai'], str(item['TenDeTai'])])
            with open('DanhSachDKDT.csv') as myfile:
                response = HttpResponse(myfile, content_type='text/csv')
                response['Content-Disposition'] = 'attachment; filename=DanhSachDKDT.csv'
                return response
        except  Exception as err:
            print(err)
@csrf_exempt
def them_thongbao(request):
    if (request.method == "GET"):
        return render(request, 'portal/admin/them_thongbao.html')
    else:
        tenHoatDong = request.POST['tenThongBao']
        Ngay = request.POST['ngayKTHD']
        ChiTiet = request.POST['chiTietHD']
        jsonRender = {'tieude' : 'Thành công', 'ThongBao' : 'Thành Công','ChiTiet' : 'Thêm thông báo thành công', 'backlink': '/admin/thongbao'}
        updatesql = "INSERT INTO `portal`.`portal_thongbao`(`ChiTiet`, `NgayThongBao`, `TieuDe`) VALUES ('{0}', '{1}', '{2}')".format(ChiTiet, Ngay,tenHoatDong)
        try:
            ChucNang.UpdateDuLieu(updatesql)
        except Exception as exc:
            jsonRender['ChiTiet'] = str(exc)
            jsonRender['ThongBao'] = str("Không thành công")
        return render(request, 'portal/giangvien/thongbao.html', jsonRender)
def xoa_thongbao(request, idTB):
    sql = "DELETE FROM `portal`.`portal_thongbao` WHERE `IdThongBao` = {0}".format(idTB)
    jsonRender = {'tieude' : 'Thành công', 'ThongBao' : 'Thành Công','ChiTiet' : 'Xóa thông báo thành công', 'backlink': '/admin/thongbao'}
    try:
        ChucNang.UpdateDuLieu(sql)
    except Exception as exc:
        jsonRender['ChiTiet'] = str(exc)
        jsonRender['ThongBao'] = str("Không thành công")
    return render(request, 'portal/giangvien/thongbao.html', jsonRender)
# 
@csrf_exempt
def loaihoatdong(request):
    if request.method == "GET":
        sql = "SELECT * FROM `portal`.`portal_loaihd`"
        dsLoaiHD = ChucNang.TruyVanDuLieu(sql)
        jsonRender = {"DsLoai":dsLoaiHD["data"]}
        # jsonRender = {'tieude' : 'Thành công', 'ThongBao' : 'KHONG CO PHEP','ChiTiet' : 'POST REQUIRED', 'backlink': '/admin/dshoatdong'}
        return render(request, 'portal/admin/ql_loaihd.html', jsonRender)
    else:
        dsSV = json.loads(request.body)
        ten = dsSV['Ten']
        resp = {"code": 200}
        resp['msg'] = "OK"
        try:
            sql ="INSERT INTO `portal`.`portal_loaihd`(`TenLoaiHD`) VALUES ('{0}')".format(ten)
            ChucNang.UpdateDuLieu(sql)
        except Exception as ere:
            resp['msg'] = str(ere)
        return HttpResponse(json.dumps(resp))
@csrf_exempt
def loaidetai(request):
    if request.method == "GET":
        sql = "SELECT * FROM `portal`.`portal_loaidetai`"
        dsLoaiHD = ChucNang.TruyVanDuLieu(sql)
        jsonRender = {"DsLoai":dsLoaiHD["data"]}
        # jsonRender = {'tieude' : 'Thành công', 'ThongBao' : 'KHONG CO PHEP','ChiTiet' : 'POST REQUIRED', 'backlink': '/admin/dshoatdong'}
        return render(request, 'portal/admin/ql_loaidt.html', jsonRender)
    else:
        dsSV = json.loads(request.body)
        ten = dsSV['Ten']
        resp = {"code": 200}
        resp['msg'] = "OK"
        try:
            sql ="INSERT INTO `portal`.`portal_loaidetai`(`TenLoai`) VALUES ('{0}')".format(ten)
            ChucNang.UpdateDuLieu(sql)
        except Exception as ere:
            resp['msg'] = str(ere)
        return HttpResponse(json.dumps(resp))
def xoaloaihoatdong(request,id):
    resp = {"code": 200}
    resp['msg'] = "success"
    if request.method == "GET":
        sql = "SELECT * FROM `portal`.`portal_hoatdong` where Loai = {0}".format(id)
        ds = ChucNang.TruyVanDuLieu(sql)
        if (len(ds['data'])==0):
            sqlDel = "DELETE FROM `portal`.`portal_loaihd` WHERE `idLoaiHD` = {0}".format(id)
            try:
                ChucNang.UpdateDuLieu(sqlDel)
                resp['msg'] = "OK"
            except  Exception as ere:
                resp = {"code": 404}
                resp['msg'] = str(ere)
            return HttpResponse(json.dumps(resp))
        resp['msg'] = "Khong the xoa, da co hoat dong dang ki"
        return HttpResponse(json.dumps(resp))
        
def xoaloaidetai(request,id):
    resp = {"code": 200}
    resp['msg'] = "success"
    if request.method == "GET":
        sql = "SELECT * FROM `portal`.`portal_detai` where IdLoai = {0}".format(id)
        ds = ChucNang.TruyVanDuLieu(sql)
        if (len(ds['data'])==0):
            sqlDel = "DELETE FROM `portal`.`portal_loaidetai` WHERE `IdLoai` = {0}".format(id)
            try:
                ChucNang.UpdateDuLieu(sqlDel)
                resp['msg'] = "OK"
            except  Exception as ere:
                resp = {"code": 404}
                resp['msg'] = str(ere)
            return HttpResponse(json.dumps(resp))
        resp['msg'] = "Khong the xoa, da co hoat dong dang ki"
        return HttpResponse(json.dumps(resp))
def exportDshddadk(request):
    if (request.method == "GET"):
        sql ="""
        SELECT
            portal_hoatdongdadangky.IdHDDDK,
            portal_hoatdongdadangky.IdHoatDong,
            portal_hoatdongdadangky.IdUser,
            portal_hoatdongdadangky.NgayDKHD,
            portal_hoatdongdadangky.DaChamDiem,
            portal_nguoidung.HoTen,
            portal_nguoidung.TenNguoiDung,
            hoatdong_giangvien.GV_username,
            hoatdong_giangvien.ID_gv,
            hoatdong_giangvien.TenGiangVien,
            portal_hoatdong.TenHoatDong, 
            portal_hoatdong.SoLuong, 
            portal_hoatdong.DangThucHien,
            portal_hoatdong.Ki,
            portal_loaihd.TenLoaiHD
            FROM
            portal_hoatdongdadangky
            JOIN portal_nguoidung ON portal_hoatdongdadangky.IdUser = portal_nguoidung.IdUser
            JOIN portal_hoatdong on portal_hoatdongdadangky.IdHoatDong = portal_hoatdong.IdHoatDong
            JOIN (SELECT  portal_hoatdong.Loai, portal_hoatdong.IdHoatDong as hdid,portal_nguoidung.IdUser as ID_gv,portal_nguoidung.HoTen as TenGiangVien, portal_nguoidung.TenNguoiDung as GV_username FROM portal_hoatdong JOIN portal_nguoidung ON portal_nguoidung.IdUser = portal_hoatdong.IdUser) AS 
            hoatdong_giangvien
            ON hoatdong_giangvien.hdid= portal_hoatdongdadangky.IdHoatDong
            JOIN portal_loaihd on hoatdong_giangvien.Loai = portal_loaihd.idLoaiHD
        """
        dataexport_res = ChucNang.TruyVanDuLieu(sql)
        header = 0
        try:
            with open('DanhSachDaDangKi.csv', mode='w',newline='', encoding='UTF-8') as csvfile:
                for item in dataexport_res['data']:
                    csvfile_writer_writer = csv.writer(csvfile, delimiter=',')
                    if (header ==0):
                        csvfile_writer_writer.writerow(['IdHDDDK','IdHoatDong','IdUser','ID_gv','GV_username','TenNguoiDung','HoTen','TenGiangVien','TenHoatDong','SoLuong','DangThucHien','Ki','TenLoaiHD','NgayDKHD','DaChamDiem'])
                        header = 1
                    csvfile_writer_writer.writerow([item['IdHDDDK'],item['IdHoatDong'],item['IdUser'],item['ID_gv'],item['GV_username'],item['TenNguoiDung'],item['HoTen'],item['TenGiangVien'],item['TenHoatDong'],item['SoLuong'],item['DangThucHien'],item['Ki'],item['TenLoaiHD'],item['NgayDKHD'],item['DaChamDiem']])
            with open('DanhSachDaDangKi.csv') as myfile:
                response = HttpResponse(myfile, content_type='text/csv')
                response['Content-Disposition'] = 'attachment; filename=DanhSachDaDangKi.csv'
                return response
        except  Exception as err:
            print(err)
def dshoatdongdk(request):
    if (request.method == "GET"):
        sql = """
            SELECT
                portal_hoatdongdadangky.IdHDDDK,
                portal_hoatdongdadangky.IdHoatDong,
                portal_hoatdongdadangky.IdUser,
                portal_hoatdongdadangky.NgayDKHD,
                portal_hoatdongdadangky.DaChamDiem,
                portal_nguoidung.HoTen,
                portal_nguoidung.TenNguoiDung,
                hoatdong_giangvien.GV_username,
                hoatdong_giangvien.ID_gv,
                hoatdong_giangvien.TenGiangVien,
                portal_hoatdong.TenHoatDong, 
                portal_hoatdong.SoLuong, 
                portal_hoatdong.DangThucHien,
                portal_hoatdong.Ki,
                portal_loaihd.TenLoaiHD
                FROM
                portal_hoatdongdadangky
                JOIN portal_nguoidung ON portal_hoatdongdadangky.IdUser = portal_nguoidung.IdUser
                JOIN portal_hoatdong on portal_hoatdongdadangky.IdHoatDong = portal_hoatdong.IdHoatDong
                JOIN (SELECT  portal_hoatdong.Loai, portal_hoatdong.IdHoatDong as hdid,portal_nguoidung.IdUser as ID_gv,portal_nguoidung.HoTen as TenGiangVien, portal_nguoidung.TenNguoiDung as GV_username FROM portal_hoatdong JOIN portal_nguoidung ON portal_nguoidung.IdUser = portal_hoatdong.IdUser) AS 
                hoatdong_giangvien
                ON hoatdong_giangvien.hdid= portal_hoatdongdadangky.IdHoatDong
                JOIN portal_loaihd on hoatdong_giangvien.Loai = portal_loaihd.idLoaiHD
            """
        data = ChucNang.TruyVanDuLieu(sql)
        jsonRender = {
            "DS_NguoiDung" :data['data']
        }
        return render(request, 'portal/admin/ql_detaidadangki.html', jsonRender) 
def dsdetai(request):
    sql = """
        SELECT
            portal_detai.IdDeTai,
            portal_detai.IdUser,
            portal_detai.ChiTiet,
            portal_detai.NgayBD,
            portal_detai.NgayKT,
            portal_detai.SoLuong,
            portal_detai.IdLoai,
            portal_detai.HoatDong,
            portal_detai.TenDeTai,
            portal_detai.DaDangKi,
            portal_detai.DangThucHien,
            portal_detai.IdKhoa,
            portal_nguoidung.HoTen,
            portal_nguoidung.TenNguoiDung,
            portal_loaidetai.TenLoai,
            portal_khoa.TenKhoa
            FROM
            portal_detai
            JOIN portal_nguoidung ON portal_detai.IdUser = portal_nguoidung.IdUser
            JOIN portal_khoa ON portal_khoa.IdKhoa = portal_detai.IdKhoa
            JOIN portal_loaidetai ON portal_loaidetai.IdLoai = portal_detai.IdLoai
            """
    data = ChucNang.TruyVanDuLieu(sql)
    jsonRender = {
        "DS_NguoiDung" :data['data']
    }
    return render(request, 'portal/admin/dsdetai.html', jsonRender) 
def dsdkdetai(request):
    sql = """
        SELECT
        portal_detaidadangky.IdDTDDK,
        portal_detaidadangky.IdUser,
        portal_detaidadangky.NgayDKDT,
        portal_nguoidung.HoTen,
        portal_nguoidung.TenNguoiDung as MSSV,
        detai_giangvien.TenGiangVien,
        detai_giangvien.IDGV,
        portal_detaidadangky.IdDeTai,
        detai_giangvien.TenDeTai

        FROM
        portal_detaidadangky
        JOIN portal_nguoidung ON portal_nguoidung.IdUser = portal_detaidadangky.IdUser
        JOIN (SELECT TenDeTai, HoTen as TenGiangVien, portal_nguoidung.IdUser as IDGV,portal_detai.IdDeTai FROM portal_detai join portal_nguoidung  on portal_nguoidung.IdUser = portal_detai.IdUser) as detai_giangvien on detai_giangvien.IdDeTai =  portal_detaidadangky.IdDeTai
            """
    data = ChucNang.TruyVanDuLieu(sql)
    jsonRender = {
        "DS_NguoiDung" :data['data']
    }
    return render(request, 'portal/admin/dsdkdetai.html', jsonRender) 
def diemrenluyen(request, ki = 0):
    sql = """SELECT
                Sum(portal_diemrenluyen.Diem) AS Diem,
                portal_nguoidung.HoTen,
                portal_diemrenluyen.Ki,
                portal_nguoidung.IdUser,
                portal_nguoidung.TenNguoiDung
                FROM
                portal_diemrenluyen
                JOIN portal_nguoidung ON portal_nguoidung.IdUser = portal_diemrenluyen.userID
                WHERE portal_nguoidung.HoatDong = 1
                GROUP BY
                portal_nguoidung.IdUser,
                portal_diemrenluyen.Ki
"""
    if (ki != 0):
        sql = """
                SELECT
                Sum(portal_diemrenluyen.Diem) AS Diem,
                portal_nguoidung.HoTen,
                portal_diemrenluyen.Ki,
                portal_nguoidung.IdUser,
                portal_nguoidung.TenNguoiDung
                FROM
                portal_diemrenluyen
                JOIN portal_nguoidung ON portal_nguoidung.IdUser = portal_diemrenluyen.userID
                WHERE
                portal_diemrenluyen.Ki = {0} AND portal_nguoidung.HoatDong = 1
                GROUP BY
                portal_nguoidung.IdUser,
                portal_diemrenluyen.Ki
        """.format(ki)
    data = ChucNang.TruyVanDuLieu(sql)
    jsonRender = {
    "DS_NguoiDung" :data['data']
    }
    return render(request, 'portal/admin/diemrenluyen.html', jsonRender) 